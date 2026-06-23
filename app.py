import os
import json
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from dotenv import load_dotenv
import openpyxl
from io import BytesIO
from flask import send_file

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "fallback_secret_key")

# ============================================================
# KONFIGURASI — isi GMAIL_APP_PASS di file .env
# ============================================================
GMAIL_SENDER   = os.getenv("GMAIL_SENDER", "fikrirestualamm004@gmail.com")
GMAIL_APP_PASS = os.getenv("GMAIL_APP_PASS", "ISI_APP_PASSWORD_DI_SINI")
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin123")

# ============================================================
# DATABASE SEMENTARA (IN-MEMORY) — data awal
# ============================================================
mahasiswa_db = [
    {"nim": "24101145056", "nama": "Andi Cobra",       "jurusan": "Sistem Informasi",   "ipk": 3.86, "email": "24101145056@kampus.ac.id", "semester": "6", "telepon": "081200000001", "tahun_masuk": "2021"},
    {"nim": "24101145072", "nama": "Giselma",           "jurusan": "Ilmu Hukum",         "ipk": 3.88, "email": "24101145072@kampus.ac.id", "semester": "6", "telepon": "081200000002", "tahun_masuk": "2021"},
    {"nim": "24101145071", "nama": "Ilham God",         "jurusan": "Teknik Informatika", "ipk": 3.98, "email": "24101145071@kampus.ac.id", "semester": "6", "telepon": "081200000003", "tahun_masuk": "2021"},
    {"nim": "24101145062", "nama": "Jefri Nichol",      "jurusan": "Pertanian",          "ipk": 3.78, "email": "24101145062@kampus.ac.id", "semester": "6", "telepon": "081200000004", "tahun_masuk": "2021"},
    {"nim": "24101145083", "nama": "Joko Anwar",        "jurusan": "DKV",                "ipk": 3.92, "email": "24101145083@kampus.ac.id", "semester": "6", "telepon": "081200000005", "tahun_masuk": "2021"},
    {"nim": "24101155042", "nama": "Windah Basuradar",  "jurusan": "Ilmu Komunikasi",    "ipk": 3.87, "email": "24101155042@kampus.ac.id", "semester": "6", "telepon": "081200000006", "tahun_masuk": "2021"},
]

# ============================================================
# HELPER — kirim email via Gmail SMTP
# ============================================================
def kirim_email_smtp(to_name, to_email, ipk, pengirim="Admin", pesan=""):
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Evaluasi Akademik - {to_name}"
        msg["From"]    = GMAIL_SENDER
        msg["To"]      = to_email

        pesan_html = pesan.replace("\n", "<br>") if pesan else ""

        html_body = f"""
        <html>
        <body style="font-family:Segoe UI,sans-serif; background:#edf4ff; padding:30px;">
          <div style="max-width:600px; margin:auto; background:white; border-radius:15px; overflow:hidden; box-shadow:0 4px 15px rgba(0,0,0,.08);">
            <div style="background:linear-gradient(90deg,#0076ff,#00c6ff); padding:25px 30px;">
              <h2 style="color:white; margin:0;">SIM Mahasiswa</h2>
              <p style="color:#e0f0ff; margin:5px 0 0;">Universitas Pamulang</p>
            </div>
            <div style="padding:30px;">
              <p>Yth. <strong>{to_name}</strong>,</p>
              <p>Berikut adalah hasil evaluasi akademik Anda:</p>
              <div style="background:#f0f7ff; border-left:4px solid #0076ff; padding:15px; border-radius:5px; margin:20px 0;">
                <p style="margin:0; font-size:18px;">IPK Anda: <strong style="color:#0076ff;">{ipk:.2f}</strong></p>
                <p style="margin:5px 0 0; color:#555;">
                  {"🏆 Cumlaude — Prestasi Luar Biasa!" if ipk >= 3.51 else ("⭐ Sangat Memuaskan" if ipk >= 3.00 else "✅ Memuaskan")}
                </p>
              </div>
              {f'''
              <div style="background:#fffbea; border-left:4px solid #ffc107; padding:15px; border-radius:5px; margin:20px 0;">
                <p style="margin:0 0 8px; font-size:13px; color:#888;">Pesan dari <strong style="color:#333;">{pengirim}</strong>:</p>
                <p style="margin:0; color:#333; line-height:1.7;">{pesan_html}</p>
              </div>
              ''' if pesan_html else ""}
              <p>Selamat atas pencapaian akademik Anda. Teruslah semangat dan pertahankan prestasi ini.</p>
              <p style="color:#888; font-size:12px; margin-top:30px;">Email ini dikirim otomatis oleh SIM Mahasiswa Universitas Pamulang.</p>
            </div>
          </div>
        </body>
        </html>
        """

        msg.attach(MIMEText(html_body, "html"))

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_SENDER, GMAIL_APP_PASS)
            server.sendmail(GMAIL_SENDER, to_email, msg.as_string())

        return True, "Email berhasil dikirim"
    except smtplib.SMTPAuthenticationError:
        return False, "Autentikasi Gmail gagal. Periksa App Password di file .env"
    except Exception as e:
        return False, str(e)


# ============================================================
# ROUTES
# ============================================================

@app.route("/")
def index():
    if "logged_in" not in session:
        return redirect(url_for("login"))
    return render_template("index.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        data = request.get_json()
        if data["username"] == ADMIN_USERNAME and data["password"] == ADMIN_PASSWORD:
            session["logged_in"] = True
            return jsonify({"success": True})
        return jsonify({"success": False, "message": "Username atau password salah!"})
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# --- API MAHASISWA ---

@app.route("/api/mahasiswa", methods=["GET"])
def get_mahasiswa():
    if "logged_in" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    return jsonify(mahasiswa_db)

@app.route("/api/mahasiswa", methods=["POST"])
def tambah_mahasiswa():
    if "logged_in" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    required = ["nim", "nama", "jurusan", "ipk", "email"]
    for field in required:
        if not data.get(field):
            return jsonify({"success": False, "message": f"Field '{field}' wajib diisi"}), 400

    # Cek NIM duplikat
    if any(m["nim"] == data["nim"] for m in mahasiswa_db):
        return jsonify({"success": False, "message": "NIM sudah terdaftar!"}), 400

    mahasiswa_db.append({
        "nim":         data["nim"],
        "nama":        data["nama"],
        "jurusan":     data["jurusan"],
        "ipk":         float(data["ipk"]),
        "email":       data["email"],
        "semester":    data.get("semester", ""),
        "telepon":     data.get("telepon", ""),
        "tahun_masuk": data.get("tahun_masuk", ""),
    })
    return jsonify({"success": True, "message": "Mahasiswa berhasil ditambahkan"})

@app.route("/api/mahasiswa/<nim>", methods=["DELETE"])
def hapus_mahasiswa(nim):
    if "logged_in" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    global mahasiswa_db
    before = len(mahasiswa_db)
    mahasiswa_db = [m for m in mahasiswa_db if m["nim"] != nim]
    if len(mahasiswa_db) < before:
        return jsonify({"success": True, "message": "Data dihapus"})
    return jsonify({"success": False, "message": "NIM tidak ditemukan"}), 404

# --- API KIRIM EMAIL ---

@app.route("/api/kirim-email", methods=["POST"])
def kirim_email():
    if "logged_in" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    data     = request.get_json()
    nama     = data.get("nama")
    email    = data.get("email")
    ipk      = float(data.get("ipk", 0))
    pengirim = data.get("pengirim", "Admin")
    pesan    = data.get("pesan", "")

    if not email:
        return jsonify({"success": False, "message": "Email mahasiswa tidak tersedia"}), 400

    ok, msg = kirim_email_smtp(nama, email, ipk, pengirim, pesan)
    return jsonify({"success": ok, "message": msg})

# --- API EXPORT EXCEL ---

@app.route("/api/export-excel")
def export_excel():
    if "logged_in" not in session:
        return redirect(url_for("login"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Mahasiswa"

    headers = ["NIM", "Nama Mahasiswa", "Jurusan", "IPK", "Semester", "Telepon", "Email"]
    ws.append(headers)

    # Style header
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="0076FF", end_color="0076FF", fill_type="solid")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    for m in mahasiswa_db:
        ws.append([m["nim"], m["nama"], m["jurusan"], m["ipk"],
                   m.get("semester",""), m.get("telepon",""), m["email"]])

    # Auto-width kolom
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Data_Mahasiswa_SIM.xlsx"
    )

# --- API IMPORT EXCEL ---

@app.route("/api/import-excel", methods=["POST"])
def import_excel():
    if "logged_in" not in session:
        return jsonify({"error": "Unauthorized"}), 401

    file = request.files.get("file")
    if not file or not file.filename.endswith(".xlsx"):
        return jsonify({"success": False, "message": "Upload file Excel (.xlsx) yang valid"}), 400

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # Ambil header dari baris pertama, normalize ke lowercase tanpa spasi
        headers = [str(cell.value or "").strip().lower().replace(" ", "_") for cell in ws[1]]

        # Mapping nama kolom fleksibel
        col_map = {}
        for i, h in enumerate(headers):
            if "nim" in h:                          col_map["nim"]         = i
            elif "nama" in h:                       col_map["nama"]        = i
            elif "jurusan" in h:                    col_map["jurusan"]     = i
            elif "ipk" in h:                        col_map["ipk"]         = i
            elif "email" in h:                      col_map["email"]       = i
            elif "semester" in h:                   col_map["semester"]    = i
            elif "telepon" in h or "telp" in h:     col_map["telepon"]     = i
            elif "tahun" in h:                      col_map["tahun_masuk"] = i

        required = ["nim", "nama", "jurusan", "ipk", "email"]
        missing  = [r for r in required if r not in col_map]
        if missing:
            return jsonify({"success": False, "message": f"Kolom wajib tidak ditemukan: {', '.join(missing)}"}), 400

        nim_existing = {m["nim"] for m in mahasiswa_db}
        ditambah = 0
        dilewati = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue  # skip baris kosong

            def get(key):
                idx = col_map.get(key)
                return str(row[idx]).strip() if idx is not None and row[idx] is not None else ""

            nim = get("nim")
            if not nim:
                continue
            if nim in nim_existing:
                dilewati += 1
                continue

            try:
                ipk_val = float(get("ipk"))
            except ValueError:
                dilewati += 1
                continue

            mahasiswa_db.append({
                "nim":         nim,
                "nama":        get("nama"),
                "jurusan":     get("jurusan"),
                "ipk":         ipk_val,
                "email":       get("email"),
                "semester":    get("semester"),
                "telepon":     get("telepon"),
                "tahun_masuk": get("tahun_masuk"),
            })
            nim_existing.add(nim)
            ditambah += 1

        pesan = f"{ditambah} data berhasil diimpor"
        if dilewati:
            pesan += f", {dilewati} dilewati (duplikat/tidak valid)"
        return jsonify({"success": True, "message": pesan})

    except Exception as e:
        return jsonify({"success": False, "message": f"Gagal membaca file: {str(e)}"}), 500


# --- API PENCARIAN (server-side) ---

@app.route("/api/cari", methods=["GET"])
def cari_mahasiswa():
    if "logged_in" not in session:
        return jsonify({"error": "Unauthorized"}), 401

    metode   = request.args.get("metode", "linear")
    kategori = request.args.get("kategori", "nama")
    keyword  = request.args.get("keyword", "").lower()
    ipk_min  = float(request.args.get("ipk_min", 0))
    ipk_max  = float(request.args.get("ipk_max", 4.0))

    hasil = []

    if metode in ("linear", "sequential"):
        hasil = [m for m in mahasiswa_db
                 if keyword in m.get(kategori, "").lower()]

    elif metode == "binary":
        data_sorted = sorted(mahasiswa_db, key=lambda m: m.get(kategori, "").lower())
        lo, hi = 0, len(data_sorted) - 1
        while lo <= hi:
            mid     = (lo + hi) // 2
            val_mid = data_sorted[mid].get(kategori, "").lower()
            if keyword in val_mid:
                hasil.append(data_sorted[mid])
                l = mid - 1
                while l >= 0 and keyword in data_sorted[l].get(kategori, "").lower():
                    hasil.append(data_sorted[l]); l -= 1
                r = mid + 1
                while r < len(data_sorted) and keyword in data_sorted[r].get(kategori, "").lower():
                    hasil.append(data_sorted[r]); r += 1
                break
            elif val_mid < keyword:
                lo = mid + 1
            else:
                hi = mid - 1

    elif metode == "rentang":
        hasil = [m for m in mahasiswa_db if ipk_min <= m["ipk"] <= ipk_max]

    return jsonify(hasil)


# ============================================================
if __name__ == "__main__":
    app.run(debug=True, port=5000)
